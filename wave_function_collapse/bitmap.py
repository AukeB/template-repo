import os
from openpyxl import load_workbook
from collections import namedtuple

Size = namedtuple('Size', ['width', 'height'])

class BitmapUtils:
    """ """
    def _obtain_bitmap_size(self, sheet):
        """Find the first column and row with None values in the given sheet."""
        first_column_with_none = None
        first_row_with_none = None

        for col_idx, cell in enumerate(sheet[1], start=1): # Access the first row
            if cell.value is None:
                first_column_with_none = col_idx - 1
                break

        for row_idx, cell in enumerate(sheet['A'], start=1): # Access the first column
            if cell.value is None:
                first_row_with_none = row_idx - 1
                break

        return first_column_with_none, first_row_with_none

    def _hex_to_rgb(self, color_hex):
        """ """       
        r = int(color_hex[0:2], 16)
        g = int(color_hex[2:4], 16)
        b = int(color_hex[4:6], 16)
        
        return (r, g, b)

    def read_bitmap_from_excel(
        self,
        file_name: str,
        relative_dir_path: str='bitmaps',
    ) -> tuple[list[list[str]], list[list[str]]]:
        """ """
        relative_file_path = os.path.join(relative_dir_path, file_name)
        workbook = load_workbook(relative_file_path)
        sheet = workbook.active
        Size_size_width, Size_size_height = self._obtain_bitmap_size(sheet=sheet)
        
        bitmap = []
        
        for row in sheet.iter_rows(min_row=1, max_row=Size_size_width, min_col=1, max_col=Size_size_height):
            bitmap_row = []

            for cell in row:
                color_hex = cell.fill.start_color.index
                color_rgb = self._hex_to_rgb(color_hex=color_hex[2:]) # First two elements contain transparency/alpha/opacity.
                bitmap_row.append(color_rgb)
            
            bitmap.append(bitmap_row)
        
        return bitmap
    
    def create_color_mapping(self, rgb_Size):
        """ """
        color_mapping = {}
        current_char = 'A'
        
        for row in rgb_Size:
            for rgb in row:
                if rgb not in color_mapping:
                    color_mapping[rgb] = current_char
                    # Fails when the maximum number of codepoints within Unicode has been 
                    # reached (somewhere at 1.1M), so that's fine.
                    current_char = chr(ord(current_char) + 1)
        
        return color_mapping
    
    def apply_color_mapping(self, rgb_Size, color_mapping):
        """ """
        return [[color_mapping[rgb] for rgb in row] for row in rgb_Size]